from openpyxl import load_workbook
from openpyxl.cell import MergedCell
import pandas as pd

def write_to_excel(template_path, output_path, table_data):
    """
    Populate tables in an Excel template based on table titles in the Excel sheet.

    Parameters:
        template_path (str): Path to the Excel template.
        output_path (str): Path to save the updated Excel file.
        table_data (dict): Dictionary where keys are table titles in the Excel template
                           and values are Pandas DataFrames to populate the tables.
    """
    # Load the Excel template
    wb = load_workbook(template_path)
    ws = wb.active  # Assuming all tables are in the active sheet

    for table_title, df in table_data.items():
        # Locate the table title in the Excel template
        title_row, title_col = None, None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == table_title:  # Match the table title
                    title_row, title_col = cell.row, cell.column
                    break
            if title_row:
                break

        if not title_row or not title_col:
            print(f"Table title '{table_title}' not found in template!")
            continue

        # Write the DataFrame to the table below the title
        start_row = title_row + 2 
        start_col = title_col

        multiindex_offset = 1
        if isinstance(df.index, pd.MultiIndex):
            multiindex_offset = len(df.index.levels)
        
        # Write DataFrame row index
        for i, idx in enumerate(df.index):
            if isinstance(idx, tuple):
                for j, jdx in enumerate(idx):
                    try:
                        ws.cell(row=start_row + i, column=start_col+j, value=jdx)
                    except:
                        pass
            else:
                try:
                    ws.cell(row=start_row + i, column=start_col, value=idx)
                except:
                    pass

        # Write DataFrame columns and values
        for i, (idx, row) in enumerate(df.iterrows()):
            for j, value in enumerate(row):
                if pd.notnull(value):
                    ws.cell(row=start_row + i, column=start_col + j + multiindex_offset, value=value)

    # Save the updated Excel file
    wb.save(output_path)
    print(f"Data written successfully to {output_path}")