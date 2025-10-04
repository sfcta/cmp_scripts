import sys, os, argparse, toml
import numpy as np
import pandas as pd
from pathlib import Path
from survey.processed_survey import ProcessedSurvey

from openpyxl import load_workbook
from openpyxl.cell import MergedCell

def write_to_excel(template_path, output_path, table_data):
    """
    Populate tables in an Excel template based on table titles in the Excel sheet.

    Parameters:
        template_path (str): Path to the Excel template.
        output_path (str): Path to save the updated Excel file.
        table_data (dict): Dictionary where keys are table titles in the Excel template
                           and values are Pandas DataFrames to populate the tables.
    """
    # Load the Excel template
    wb = load_workbook(template_path)
    ws = wb.active  # Assuming all tables are in the active sheet

    for table_title, df in table_data.items():
        # Locate the table title in the Excel template
        title_row, title_col = None, None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == table_title:  # Match the table title
                    title_row, title_col = cell.row, cell.column
                    break
            if title_row:
                break

        if not title_row or not title_col:
            print(f"Table title '{table_title}' not found in template!")
            continue

        # Write the DataFrame to the table below the title
        start_row = title_row + 2 
        start_col = title_col

        merged_cell_offset = 1
        while isinstance(ws.cell(row=start_row, column=start_col+merged_cell_offset), MergedCell):
            merged_cell_offset += 1

        # Write DataFrame row index
        for i, idx in enumerate(df.index):
            ws.cell(row=start_row + i, column=start_col, value=idx)

        # Write DataFrame columns and values
        for i, (idx, row) in enumerate(df.iterrows()):
            for j, value in enumerate(row):
                ws.cell(row=start_row + i, column=start_col + j + merged_cell_offset, value=value)

    # Save the updated Excel file
    wb.save(output_path)
    print(f"Data written successfully to {output_path}")

def read_taz_to_agg_geo(config):
    file = config['taz_to_agg_geo']['file']
    taz_field = config['taz_to_agg_geo']['taz_field']
    geo_field = config['taz_to_agg_geo']['geo_field']
    taz_to_agg_geo = (pd.read_csv(file)[[taz_field, geo_field]]
                      .rename(columns={taz_field:'taz',
                                       geo_field:'geo'})
                     )
    return taz_to_agg_geo

def get_trips_with_cmp_mode_type(survey, config):
    '''
    create new `cmp_mode`
    1 = walk
    2 = bike
    3 = drive alone
    4 = shared ride 2
    5 = shared ride 3+
    6 = walk transit
    7 = drive transit
    8 = school bus
    9 = tnc
    10 = scooter share
    11 = bike share
    
    simply copies daysim modes, except where daysim mode is walk (1) or bike (2) and raw survey mode_type is bikeshare (3) or scootershare (4)
    '''
    weight = config['survey']['trip']['weight']
    trip = pd.merge(survey.trip.data[['hhno','pno','tsvid','otaz','dtaz','mode',weight]],
                    survey._raw_trip.data[['hhno','pno','tsvid','mode_type']])

    trip['cmp_mode'] = trip['mode']
    trip.loc[trip['mode'].isin([7,8]), 'cmp_mode'] = 6
    trip.loc[trip['mode'].isin([1,2]) & trip['mode_type'].eq(3),'cmp_mode'] = 11 # bike share
    trip.loc[trip['mode'].isin([1,2]) & trip['mode_type'].eq(4),'cmp_mode'] = 10 # scooter share

    return trip

def attach_aggregate_od_geo(trip, taz_to_agg_geo):
    trip = pd.merge(trip, taz_to_agg_geo.rename(columns={'taz':'otaz','geo':'ogeo'}), how='left')
    trip = pd.merge(trip, taz_to_agg_geo.rename(columns={'taz':'dtaz','geo':'dgeo'}), how='left')
    return trip

def od_modesum(trip, ogeo, dgeo, how, mode_field, weight_field):
    sum = od_modetotal(trip, ogeo, dgeo, how, mode_field, weight_field)
    return sum.divide(sum.sum())
    
def od_modetotal(trip, ogeo, dgeo, how, mode_field, weight_field):
    how = how.lower()
    
    if how == 'or':
        t = trip.loc[trip['ogeo'].eq(ogeo) | trip['dgeo'].eq(dgeo)]
    elif how == 'xor':
        t = trip.loc[(trip['ogeo'].eq(ogeo) & trip['dgeo'].ne(dgeo)) |
                     (trip['dgeo'].eq(dgeo) & trip['ogeo'].ne(ogeo))]
    elif how == 'and':
        t = trip.loc[trip['ogeo'].eq(ogeo) & trip['dgeo'].eq(dgeo)]
    else:
        raise Exception('unrecognized `how` {}'.format(how))
    return t.groupby(mode_field).agg({weight_field:'sum'})
    
def run_modeshare(config):
    # get config settings
    survey_args = config['survey']
    outfile = Path(config['output']['dir']) / config['output']['file']
    cmp_mode_field = 'cmp_mode'
    cmp_mode_name_field = 'cmp_mode_name'
    weight_field = config['survey']['trip']['weight']
    template = Path(config['template']['dir']) / config['template']['file']
    
    # set up output dataframes
    mi = pd.MultiIndex.from_tuples([(3, 'drive alone'),
                                (9, 'tnc'),
                                (4, 'shared ride 2'),
                                (5, 'shared ride 3+'),
                                (6, 'transit'),
                                (1, 'walk'),
                                (2, 'bike'),
                                (11, 'bike share'),
                                (10, 'scooter share')], 
                               names=[cmp_mode_field,cmp_mode_name_field])
    tot = pd.DataFrame(index=mi, columns=config['modesums'].keys())
    tot.reset_index(level=1, inplace=True)
    shr = tot.copy()
    
    # read data
    survey = ProcessedSurvey(**survey_args)
    taz_to_agg_geo = read_taz_to_agg_geo(config)
    
    # preprocess data
    trip = get_trips_with_cmp_mode_type(survey, config)
    trip = attach_aggregate_od_geo(trip, taz_to_agg_geo)
    
    # prepare modesums
    for name, args in config['modesums'].items():
        # totals
        tmp = od_modetotal(trip.loc[trip['mode'].ne(0)], args['ogeo'], args['dgeo'], args['how'], cmp_mode_field, weight_field)
        tmp.rename(columns={weight_field:name}, inplace=True)
        tot.update(tmp)
        
        # shares
        tmp = od_modesum(trip.loc[trip['mode'].ne(0)], args['ogeo'], args['dgeo'], args['how'], cmp_mode_field, weight_field)
        tmp.rename(columns={weight_field:name}, inplace=True)
        shr.update(tmp)

        table_data = {'Mode Totals':tot,
                      'Mode Shares':shr}
    
    # write outputs
    write_to_excel(template, outfile, table_data)

if __name__=='__main__':
    parser = argparse.ArgumentParser(description="Process TOML configuration file for validation.")
    parser.add_argument("config_path", type=str, help="Path to the TOML configuration file.")
    
    # Check if the script is running in an interactive environment or not
    if len(sys.argv) > 1:
        # Running from command line, use the provided argument
        args = parser.parse_args()
        
        # Load the TOML configuration file
        config = toml.load(args.config_path)

        # Run the validation function with the loaded configuration
        run_modeshare(config)
        
    else:
        print("Please supply a config file.")